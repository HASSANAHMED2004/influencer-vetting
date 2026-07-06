"""Reading rows from and writing results back to the Excel workbook.

The I/O edge of the pipeline — deliberately separate from the logic so the logic
stays unit-testable.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from . import config
from .models import RowResult, Verdict


@dataclass
class SheetRow:
    """A single influencer row as read from the sheet (raw cell values)."""

    row_index: int
    youtube: object
    instagram: object
    tiktok: object
    country: object


def _require_columns(worksheet, path: Path) -> None:
    """Fail loudly if the sheet doesn't have the layout we assume."""
    header = worksheet.cell(config.HEADER_ROW, config.COL_INSTAGRAM).value
    if header is None or "instagram" not in str(header).strip().lower():
        raise ValueError(
            f"{path}: expected an 'Instagram' header in column "
            f"{get_column_letter(config.COL_INSTAGRAM)} but found {header!r}. "
            "Check SHEET_NAME / column indices in config.py."
        )


def read_rows(path: str | Path, start_row: int, end_row: int | None) -> list[SheetRow]:
    """Read rows [start_row, end_row] inclusive (1-based, data rows only)."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"input workbook not found: {path}")

    workbook = load_workbook(path, data_only=True, read_only=True)
    if config.SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"{path}: sheet {config.SHEET_NAME!r} not found; "
                         f"available: {workbook.sheetnames}")
    worksheet = workbook[config.SHEET_NAME]
    _require_columns(worksheet, path)

    start = max(start_row, config.FIRST_DATA_ROW)
    end = end_row if end_row is not None else worksheet.max_row
    rows: list[SheetRow] = []
    for r in range(start, end + 1):
        rows.append(SheetRow(
            row_index=r,
            youtube=worksheet.cell(r, config.COL_YOUTUBE).value,
            instagram=worksheet.cell(r, config.COL_INSTAGRAM).value,
            tiktok=worksheet.cell(r, config.COL_TIKTOK).value,
            country=worksheet.cell(r, config.COL_COUNTRY).value,
        ))
    workbook.close()
    return rows


def write_results(input_path: str | Path, output_path: str | Path,
                  results: list[RowResult]) -> None:
    """Write per-row results into new columns of a copy of the workbook.

    Reads the *full* workbook (not read-only) so untouched rows are preserved,
    writes output headers + values, and sets column A True where overall == PASS.
    """
    input_path, output_path = Path(input_path), Path(output_path)
    workbook = load_workbook(input_path)  # values-and-formatting preserving copy
    worksheet = workbook[config.SHEET_NAME]

    for col, header in config.OUTPUT_HEADERS.items():
        worksheet.cell(config.HEADER_ROW, col, header)

    for res in results:
        r = res.row_index
        worksheet.cell(r, config.COL_OUT_COUNTRY_STATUS, res.country_status)
        worksheet.cell(r, config.COL_OUT_YT_MAX_VIEWS, _num(res.youtube.avg_views))
        worksheet.cell(r, config.COL_OUT_YT_VERDICT, res.youtube.verdict.value)
        worksheet.cell(r, config.COL_OUT_IG_HANDLE, res.instagram.handle)
        worksheet.cell(r, config.COL_OUT_IG_FOLLOWERS, res.instagram.followers)
        worksheet.cell(r, config.COL_OUT_IG_AVG_VIEWS, _num(res.instagram.avg_views))
        worksheet.cell(r, config.COL_OUT_IG_VERDICT, res.instagram.verdict.value)
        worksheet.cell(r, config.COL_OUT_TT_HANDLE, res.tiktok.handle)
        worksheet.cell(r, config.COL_OUT_TT_FOLLOWERS, res.tiktok.followers)
        worksheet.cell(r, config.COL_OUT_TT_AVG_VIEWS, _num(res.tiktok.avg_views))
        worksheet.cell(r, config.COL_OUT_TT_VERDICT, res.tiktok.verdict.value)
        worksheet.cell(r, config.COL_OUT_OVERALL, res.overall.value)
        worksheet.cell(r, config.COL_OUT_NOTES, "; ".join(res.notes))
        # Tick the checkbox only on a clear PASS; leave the rest for manual review.
        worksheet.cell(r, config.COL_CHECKBOX, res.overall is Verdict.PASS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _num(value: float | None) -> int | float | None:
    """Store whole numbers as ints for a cleaner sheet."""
    if value is None:
        return None
    return int(value) if float(value).is_integer() else value
