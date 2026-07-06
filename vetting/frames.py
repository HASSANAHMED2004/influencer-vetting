"""DataFrame-driven pipeline for the Streamlit UI.

Resolves the required columns by header name (so extra columns don't matter and
positions can shift), runs the per-row vetting, and turns results into row colors.
The rule: only *passing* rows are highlighted, whole-row, by which platform passed.
"""

from __future__ import annotations

from collections.abc import Callable, Collection, Sequence
from dataclasses import dataclass
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from .models import ExclusionRule, RowResult, Verdict
from .pipeline import process_row

# Required logical columns -> keywords matched against the uploaded headers
# (case-insensitive: exact match preferred, else substring).
REQUIRED_COLUMNS: dict[str, str] = {
    "youtube": "youtube",
    "instagram": "instagram",
    "tiktok": "tiktok",
    "linkedin": "linkedin",
    "country": "country",
}


@dataclass(frozen=True)
class RowStyle:
    """A whole-row highlight: a label plus fill and font colors (#RRGGBB)."""

    label: str
    fill: str
    font: str


# Per-platform colors (single platform passed) and the multi-platform color.
PLATFORM_STYLES: dict[str, RowStyle] = {
    "instagram": RowStyle("Instagram", "#800080", "#FFFFFF"),  # purple
    "youtube": RowStyle("YouTube", "#FF0000", "#FFFFFF"),  # red
    "linkedin": RowStyle("LinkedIn", "#0000FF", "#FFFFFF"),  # blue
    "tiktok": RowStyle("TikTok", "#000000", "#FFFFFF"),  # black
}
MULTI_STYLE = RowStyle("Multiple platforms", "#FFFF00", "#000000")  # yellow


class MissingColumnsError(ValueError):
    """Raised when the uploaded dataset lacks one of the required columns."""


def resolve_columns(df: pd.DataFrame) -> dict[str, str]:
    """Map each logical column to the actual header in ``df``.

    Extra columns are ignored; only the required ones must be present. Raises
    MissingColumnsError listing any that couldn't be found.
    """
    by_lower = {str(c).strip().lower(): c for c in df.columns}
    resolved: dict[str, str] = {}
    missing: list[str] = []
    for logical, keyword in REQUIRED_COLUMNS.items():
        if keyword in by_lower:
            resolved[logical] = by_lower[keyword]
            continue
        match = next((actual for low, actual in by_lower.items() if keyword in low), None)
        if match is None:
            missing.append(logical)
        else:
            resolved[logical] = match
    if missing:
        raise MissingColumnsError(
            f"Uploaded file is missing required column(s): {missing}. "
            f"Found: {list(df.columns)}"
        )
    return resolved


@dataclass
class _DFRow:
    """Adapter exposing the attributes ``process_row`` expects from a df row."""

    row_index: int
    youtube: object
    instagram: object
    tiktok: object
    linkedin: object
    country: object


def _cell(value: object) -> object:
    """Normalize pandas empties (NaN/NaT) to None so downstream logic is clean."""
    return None if pd.isna(value) else value


def run_over_dataframe(
    df: pd.DataFrame,
    colmap: dict[str, str],
    *,
    youtube_client=None,
    social_client=None,
    exclusion_rules: Sequence[ExclusionRule] | None = None,
    disabled_platforms: Collection[str] = (),
    progress: Callable[[float], None] | None = None,
) -> list[RowResult]:
    """Run the vetting pipeline over every row, in order. Returns aligned results."""
    results: list[RowResult] = []
    total = max(len(df), 1)
    for done, (position, row) in enumerate(df.iterrows(), start=1):
        adapter = _DFRow(
            row_index=int(position),  # original position in the full dataset
            youtube=_cell(row[colmap["youtube"]]),
            instagram=_cell(row[colmap["instagram"]]),
            tiktok=_cell(row[colmap["tiktok"]]),
            linkedin=_cell(row[colmap["linkedin"]]),
            country=_cell(row[colmap["country"]]),
        )
        results.append(process_row(
            adapter, youtube_client=youtube_client,
            social_client=social_client, exclusion_rules=exclusion_rules,
            disabled_platforms=disabled_platforms,
        ))
        if progress is not None:
            progress(done / total)
    return results


def passed_platforms(result: RowResult) -> list[str]:
    """Platforms whose verdict is PASS for this row."""
    pairs = (
        ("youtube", result.youtube),
        ("instagram", result.instagram),
        ("tiktok", result.tiktok),
        ("linkedin", result.linkedin),
    )
    return [name for name, pr in pairs if pr.verdict is Verdict.PASS]


def row_style(result: RowResult) -> RowStyle | None:
    """The highlight for a row: platform color if exactly one passed, yellow if
    more than one, and None (no highlight) if none passed."""
    passed = passed_platforms(result)
    if not passed:
        return None
    if len(passed) >= 2:
        return MULTI_STYLE
    return PLATFORM_STYLES[passed[0]]


def passing_frame(df: pd.DataFrame, results: list[RowResult]) -> pd.io.formats.style.Styler:
    """Return a Styler of *only* the passing rows, each filled with its color.

    Rows are matched to ``df`` by their original position (``result.row_index``)
    and kept in result order. Non-passing rows are dropped entirely. A leading
    "Row" column shows the original spreadsheet row (header is row 1, so the
    first data row is 2).
    """
    fills = {res.row_index: s for res in results if (s := row_style(res)) is not None}
    index = [i for i in df.index if i in fills]
    subset = df.loc[index].copy()
    subset.insert(0, "Row", [i + 2 for i in subset.index])  # original Excel row

    def _apply(row: pd.Series) -> list[str]:
        style = fills[row.name]
        return [f"background-color: {style.fill}; color: {style.font};"] * len(row)

    return subset.style.apply(_apply, axis=1)


def write_passing_xlsx(uploaded_bytes: bytes, results: list[RowResult]) -> bytes:
    """Build a new workbook containing only the passing rows, color-filled.

    Copies the header and each passing row's original cell values (preserving
    numbers/text) from the uploaded sheet into a fresh "Passing" sheet. Rows are
    written in result order and colored by which platform qualified them.
    """
    source = load_workbook(BytesIO(uploaded_bytes))
    src_ws = source[source.sheetnames[0]]
    ncols = src_ws.max_column

    out = Workbook()
    ws = out.active
    ws.title = "Passing"
    ws.cell(1, 1, "Row")  # original spreadsheet row number
    for col in range(1, ncols + 1):  # copy header after the Row column
        ws.cell(1, col + 1, src_ws.cell(1, col).value)

    out_row = 2
    for result in results:
        style = row_style(result)
        if style is None:
            continue
        fill = PatternFill("solid", fgColor=style.fill.lstrip("#"))
        font = Font(color=style.font.lstrip("#"))
        src_row = result.row_index + 2  # +1 header, +1 for 1-based rows
        row_cell = ws.cell(out_row, 1, src_row)
        row_cell.fill = fill
        row_cell.font = font
        for col in range(1, ncols + 1):
            cell = ws.cell(out_row, col + 1, src_ws.cell(src_row, col).value)
            cell.fill = fill
            cell.font = font
        out_row += 1

    buffer = BytesIO()
    out.save(buffer)
    return buffer.getvalue()


def color_summary(results: list[RowResult]) -> dict[str, int]:
    """Count rows per highlight label (for the UI summary)."""
    counts: dict[str, int] = {}
    for result in results:
        style = row_style(result)
        if style is None:
            continue
        label = MULTI_STYLE.label if style is MULTI_STYLE else style.label
        counts[label] = counts.get(label, 0) + 1
    return counts
