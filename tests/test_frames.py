"""Tests for the DataFrame pipeline and pass-row coloring used by the UI."""

from __future__ import annotations

import pandas as pd
import pytest

from vetting.frames import (
    MULTI_STYLE,
    PLATFORM_STYLES,
    MissingColumnsError,
    is_excluded_column,
    passing_frame,
    resolve_columns,
    row_style,
    run_over_dataframe,
    write_passing_xlsx,
)
from vetting.models import PlatformResult, RowResult, Verdict


def test_resolve_columns_by_header_ignoring_extras_and_order():
    df = pd.DataFrame(columns=[
        "Email", "TikTok", "YouTube", "LinkedIn", "Your country/region",
        "Instagram", "Random Extra",
    ])
    colmap = resolve_columns(df)
    assert colmap == {
        "youtube": "YouTube", "instagram": "Instagram", "tiktok": "TikTok",
        "linkedin": "LinkedIn", "country": "Your country/region",
    }


def test_resolve_columns_missing_raises():
    df = pd.DataFrame(columns=["YouTube", "Instagram"])  # no tiktok/country
    with pytest.raises(MissingColumnsError) as exc:
        resolve_columns(df)
    assert "tiktok" in str(exc.value) and "country" in str(exc.value)


def _result(index: int = 0, **verdicts) -> RowResult:
    res = RowResult(row_index=index)
    for platform, verdict in verdicts.items():
        setattr(res, platform, PlatformResult(verdict=verdict))
    return res


def test_row_style_single_platform_colors():
    assert row_style(_result(instagram=Verdict.PASS)) is PLATFORM_STYLES["instagram"]
    assert row_style(_result(youtube=Verdict.PASS)) is PLATFORM_STYLES["youtube"]
    assert row_style(_result(tiktok=Verdict.PASS)) is PLATFORM_STYLES["tiktok"]
    assert row_style(_result(linkedin=Verdict.PASS)) is PLATFORM_STYLES["linkedin"]


def test_row_style_multi_is_yellow():
    style = row_style(_result(youtube=Verdict.PASS, instagram=Verdict.PASS))
    assert style is MULTI_STYLE


def test_row_style_none_when_no_pass():
    assert row_style(_result(instagram=Verdict.FAIL, youtube=Verdict.REVIEW)) is None
    assert row_style(_result()) is None


def test_run_over_dataframe_free_mode_and_exclusion():
    df = pd.DataFrame({
        "YouTube": ["chan", None],
        "Instagram": ["someone", "other"],
        "TikTok": [None, None],
        "LinkedIn": [None, None],
        "Your country/region": ["Spain", "India"],  # 2nd row excluded
    })
    colmap = resolve_columns(df)
    results = run_over_dataframe(df, colmap)  # no clients -> nothing passes
    assert len(results) == 2
    assert results[1].overall is Verdict.EXCLUDED
    # Nothing passes in free mode, so no row is highlighted.
    assert all(row_style(r) is None for r in results)


def test_is_excluded_column():
    assert is_excluded_column("Discord")
    assert is_excluded_column("  discord ")  # trimmed + case-insensitive
    assert is_excluded_column("Token")
    assert is_excluded_column("Unnamed: 0")
    assert is_excluded_column("Unnamed: 7")
    assert not is_excluded_column("X")  # X is kept
    assert not is_excluded_column("Instagram")
    assert not is_excluded_column("Email")


def test_excluded_columns_dropped_from_display_and_export():
    from io import BytesIO

    from openpyxl import load_workbook

    df = pd.DataFrame({
        "Unnamed: 0": [0, 1],
        "Email": ["a@x.com", "b@x.com"],
        "Discord": ["a#1", "b#2"],
        "Instagram": ["ig_a", "ig_b"],
        "X": ["xa", "xb"],  # kept
    })
    results = [_result(0, instagram=Verdict.PASS), _result(1, instagram=Verdict.FAIL)]

    shown = passing_frame(df.astype("string").fillna(""), results).data
    assert "Discord" not in shown.columns
    assert "Unnamed: 0" not in shown.columns
    assert "Email" in shown.columns and "Instagram" in shown.columns and "X" in shown.columns

    buf = BytesIO()
    df.to_excel(buf, index=False)
    ws = load_workbook(BytesIO(write_passing_xlsx(buf.getvalue(), results))).active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert headers[:2] == ["Approved", "Row"]
    assert "Discord" not in headers and "Unnamed: 0" not in headers
    assert "Email" in headers and "Instagram" in headers and "X" in headers


def test_partial_run_preserves_original_positions():
    df = pd.DataFrame({
        "YouTube": [None, None, None],
        "Instagram": ["a", "b", "c"],
        "TikTok": [None, None, None],
        "LinkedIn": [None, None, None],
        "Your country/region": ["Spain", "Spain", "Spain"],
    })
    colmap = resolve_columns(df)
    # Vet only rows at positions 1 and 2 (a partial batch).
    results = run_over_dataframe(df.iloc[1:3], colmap)
    assert [r.row_index for r in results] == [1, 2]


def test_passing_frame_keeps_only_passing_rows_in_order():
    df = pd.DataFrame({"Name": ["Alice", "Bob", "Cara", "Dan"]})
    results = [
        _result(0, instagram=Verdict.PASS),                    # kept (purple)
        _result(1, instagram=Verdict.FAIL),                    # dropped
        _result(2, instagram=Verdict.PASS, youtube=Verdict.PASS),  # kept (yellow)
        _result(3, youtube=Verdict.PASS),                      # kept (red)
    ]
    styler = passing_frame(df, results)
    assert list(styler.data.index) == [0, 2, 3]
    assert list(styler.data["Name"]) == ["Alice", "Cara", "Dan"]
    # Leading "Row" column = original spreadsheet row (0-based pos + 2).
    assert list(styler.data["Row"]) == [2, 4, 5]
    # Leading columns: "Approved" (link) then "Row".
    assert list(styler.data.columns)[:2] == ["Approved", "Row"]


def test_passing_frame_empty_when_nothing_passes():
    df = pd.DataFrame({"Name": ["Alice", "Bob"]})
    results = [_result(0, instagram=Verdict.FAIL), _result(1, youtube=Verdict.FAIL)]
    assert len(passing_frame(df, results).data) == 0


def test_write_passing_xlsx_only_passing_rows_with_values_and_color():
    from io import BytesIO

    from openpyxl import load_workbook

    df = pd.DataFrame({
        "YouTube": ["a", "b", "c"],
        "Instagram": ["x", "y", "z"],
        "TikTok": [None, None, None],
        "LinkedIn": [None, None, None],
        "Your country/region": ["Spain", "Spain", "Spain"],
    })
    buf = BytesIO()
    df.to_excel(buf, index=False)

    results = [
        _result(0, instagram=Verdict.PASS),
        _result(1, instagram=Verdict.FAIL),  # dropped
        _result(2, youtube=Verdict.PASS),
    ]
    ws = load_workbook(BytesIO(write_passing_xlsx(buf.getvalue(), results))).active
    # Header + exactly the 2 passing rows, in order.
    assert ws.max_row == 3
    # Col 1 = Approved (platform label), col 2 = original Excel row, col 3 = first data col.
    assert [ws.cell(r, 1).value for r in (1, 2, 3)] == ["Approved", "Instagram", "YouTube"]
    assert [ws.cell(r, 2).value for r in (1, 2, 3)] == ["Row", 2, 4]
    assert [ws.cell(r, 3).value for r in (1, 2, 3)] == ["YouTube", "a", "c"]
    assert ws.cell(2, 2).fill.fgColor.rgb.endswith("800080")  # IG purple (Row cell)
    assert ws.cell(3, 3).fill.fgColor.rgb.endswith("FF0000")  # YouTube red
